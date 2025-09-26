
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

archivo_excel = 'datos.xlsx'
# comprobar si el archivo existe
if os.path.exists(archivo_excel):
    wb = load_workbook(archivo_excel)
    ws = wb.active
else:
# crear el libro de excel
    wb = Workbook()
    ws = wb.active 
    ws.append(["nombre", "edad", "email", "telefono", "direccion"])

def guardar_datos():
    nombre = entry_nombre.get()
    print(nombre)
    edad = entry_edad.get()
    print(edad)
    email = entry_email.get()
    print(email)
    telefono = entry_telefono.get()
    print(telefono)
    direccion = entry_direccion.get()
    print(direccion)

    if not nombre or not edad or not email or not telefono or not direccion:
        messagebox.showwarning("Advertencia", "Todos los campos son obligatorios")
        return
    
    try:
        edad = int(edad)
        telefono = int(telefono)
    except ValueError:
        messagebox.showwarning("Advertencia", "Edad y telefono deben de ser numeros")
        return
    
    # Se guardan los datos
    ws.append([nombre, edad, email, telefono, direccion])
    wb.save(archivo_excel)
    messagebox.showinfo("Los datos","Se guardaron correctamente")

    # Limpiar cajas de texto
    entry_nombre.delete(0,'end')
    entry_edad.delete(0,'end')
    entry_email.delete(0,'end')
    entry_telefono.delete(0,'end')
    entry_direccion.delete(0,'end')

root = tk.Tk()
root.title('Tesji')
root.configure(bg = '#4b6587')
label_style = {"bg": '#4b6587', "fg": "white"}
entry_style = {"bg": '#d3d3d3', "fg": "black"}

label_nombre = tk.Label(root, text="Nombre: ", **label_style)
label_nombre.grid(row=0, column=0, padx=10,pady=5)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady=5)

label_edad = tk.Label(root, text="Edad: ", **label_style)
label_edad.grid(row=1, column=0, padx=10,pady=5)
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=1, column=1, padx=10, pady=5)

label_email = tk.Label(root, text="Email: ", **label_style)
label_email.grid(row=2, column=0, padx=10,pady=5)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=2, column=1, padx=10, pady=5)

label_telefono = tk.Label(root, text="Telefono: ", **label_style)
label_telefono.grid(row=3, column=0, padx=10,pady=5)
entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=3, column=1, padx=10, pady=5)

label_direccion = tk.Label(root, text="Direccion: ", **label_style)
label_direccion.grid(row=4, column=0, padx=10,pady=5)
entry_direccion = tk.Entry(root, **entry_style)
entry_direccion.grid(row=4, column=1, padx=10, pady=5)

boton_guardar = tk.Button(root, text="Guardar", command=guardar_datos,bg = '#6d8299', fg='white')
boton_guardar.grid(row=5, column=0, columnspan=2, padx=10, pady=10)

root.mainloop()