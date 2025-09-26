import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
import os

# -------- CONFIGURACIÓN DE EXCEL --------
archivo_excel = 'provedores.xlsx'
if os.path.exists(archivo_excel):
    wb = load_workbook(archivo_excel)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Id.Provedor", "Nombre", "A.paterno", "A.materno", "Edad", "Direccion", "Empresa-proce", "Telefono", "Email", "Observacion"])

# -------- FUNCIÓN PARA BARRA DE PROGRESO --------
def completar_barra():
    valor = barra_progreso['value']
    if valor < 100:
        barra_progreso['value'] = valor + 10
        actualizar_porcentaje()
        root.after(100, completar_barra)
    else:
        porcentaje.set('100%')
        messagebox.showinfo("Éxito", "Los datos se guardaron correctamente")
        barra_progreso['value'] = 0
        porcentaje.set('0%')

def actualizar_porcentaje():
    valor = barra_progreso['value']
    porcentaje.set(f'{int(valor)}%')

# -------- FUNCIÓN PARA GUARDAR DATOS --------
def guardar_datos():
    idProvedor = entry_idProvedor.get()
    nombre = entry_nombre.get()
    aPaterno = entry_aPaterno.get()
    aMaterno = entry_aMaterno.get()
    edad = entry_edad.get()
    direccion = entry_direccion.get()
    empresa = entry_empresa.get()
    telefono = entry_telefono.get()
    email = entry_email.get()
    observacion = entry_observacion.get()

    if not idProvedor or not nombre or not aPaterno or not aMaterno or not edad or not direccion or not empresa or not telefono or not email or not observacion:
        messagebox.showwarning("Advertencia", "Todos los campos son obligatorios")
        return
    try:
        edad = int(edad)
        telefono = int(telefono)
        idProvedor = int(idProvedor)
    except ValueError:
        messagebox.showwarning("Advertencia", "Id de Provedor, Edad y Telefono deben ser números")
        return

    ws.append([idProvedor, nombre, aPaterno, aMaterno, edad, direccion, empresa, telefono, email, observacion])
    wb.save(archivo_excel)

    for campo in [entry_idProvedor, entry_nombre, entry_aPaterno, entry_aMaterno, entry_edad, entry_direccion, entry_empresa, entry_telefono, entry_email, entry_observacion]:
        campo.delete(0, 'end')

    barra_progreso['value'] = 0
    porcentaje.set('0%')
    completar_barra()

# -------- INTERFAZ --------
root = tk.Tk()
root.title('Formulario de Proveedores')
root.configure(bg='#4b6587')

label_style = {"bg": '#4b6587', "fg": "white"}
entry_style = {"bg": '#d3d3d3', "fg": "black"}

campos = [
    ("Id de provedor", 'entry_idProvedor'),
    ("Nombre", 'entry_nombre'),
    ("A.paterno", 'entry_aPaterno'),
    ("A.materno", 'entry_aMaterno'),
    ("Edad", 'entry_edad'),
    ("Direccion", 'entry_direccion'),
    ("Empresa de procedencia", 'entry_empresa'),
    ("Telefono", 'entry_telefono'),
    ("Email", 'entry_email'),
    ("Observacion", 'entry_observacion')
]

entradas = {}
for i, (texto, var) in enumerate(campos):
    tk.Label(root, text=texto + ":", **label_style).grid(row=i, column=0, padx=10, pady=5, sticky='w')
    entradas[var] = tk.Entry(root, **entry_style)
    entradas[var].grid(row=i, column=1, padx=10, pady=5)

entry_idProvedor = entradas['entry_idProvedor']
entry_nombre = entradas['entry_nombre']
entry_aPaterno = entradas['entry_aPaterno']
entry_aMaterno = entradas['entry_aMaterno']
entry_edad = entradas['entry_edad']
entry_direccion = entradas['entry_direccion']
entry_empresa = entradas['entry_empresa']
entry_telefono = entradas['entry_telefono']
entry_email = entradas['entry_email']
entry_observacion = entradas['entry_observacion']

# ESTILOS ttk
style = ttk.Style()
style.theme_use('clam')
style.configure('TProgressbar', troughcolor='#34495E', background='#1ABC9C', thickness=30)
style.configure('TButton', font=('helvetica', 10, 'bold'), background='#2980B9', foreground='white')
style.map('TButton', background=[('active', '#5A8CAD')])

# BARRA DE PROGRESO
barra_progreso = ttk.Progressbar(root, orient='horizontal', length=200, mode='determinate', style='TProgressbar')
barra_progreso.grid(row=len(campos), column=0, columnspan=2, pady=10)

porcentaje = tk.StringVar()
porcentaje.set('0%')
label_porcentaje = tk.Label(root, textvariable=porcentaje, font=('helvetica', 10, 'bold'), bg='#2C3E50', fg='white')
label_porcentaje.grid(row=len(campos)+1, column=0, columnspan=2)

# BOTÓN GUARDAR
boton_guardar = ttk.Button(root, text="Guardar", command=guardar_datos, style='TButton')
boton_guardar.grid(row=len(campos)+2, column=0, columnspan=2, pady=10)

root.mainloop()